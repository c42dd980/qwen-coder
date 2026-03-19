"""
PDF Search Engine - High-performance PDF indexing and search tool with OCR capabilities.
Optimized for minimal I/O, memory efficiency, and parallel throughput.
"""

import hashlib
import io
import logging
import logging.handlers
import multiprocessing
import os
import re
import subprocess
import sys
import tempfile
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

# Импорты на уровне модуля — один раз при старте
import fitz  # PyMuPDF
import pytesseract
import yaml
from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter
from plyer import notification
from tqdm import tqdm
from whoosh import index
from whoosh.fields import ID, NUMERIC, TEXT, Schema  # <-- добавлен NUMERIC
from whoosh.qparser import QueryParser


class ConfigError(Exception):
    """Exception for configuration errors."""

    pass


def setup_logging(
    log_file: str = "data/logs/",
    max_bytes: int = 10 * 1024 * 1024,
    backup_count: int = 5,
) -> logging.Logger:
    import time

    # Создаем имя файла с Unix timestamp
    timestamp = str(int(time.time()))
    log_filename = f"{log_file}{timestamp}.log"

    # Создаем директорию, если она не существует
    Path(log_filename).parent.mkdir(parents=True, exist_ok=True)

    file_handler = logging.handlers.RotatingFileHandler(
        log_filename, maxBytes=max_bytes, backupCount=backup_count, encoding="utf-8"
    )
    file_formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
    )
    file_handler.setFormatter(file_formatter)

    console_handler = logging.StreamHandler()
    console_formatter = logging.Formatter("%(levelname)s: %(message)s")
    console_handler.setFormatter(console_formatter)
    console_handler.setLevel(logging.WARNING)

    logging.basicConfig(
        level=logging.INFO, handlers=[file_handler, console_handler], force=True
    )
    logger = logging.getLogger(__name__)

    # Сохраняем путь к лог-файлу в глобальную переменную для доступа из других частей кода
    global LOG_FILE_PATH
    LOG_FILE_PATH = log_filename

    return logger


class ConsoleProgressBar:
    def __init__(self, title: str = "Processing ", total: int = 100):
        self.pbar = tqdm(
            total=total,
            desc=title,
            unit="it",
            ncols=100,
            colour="green",
            leave=False,
        )

    def update(self, value: float, text: str = ""):
        increment = value - self.pbar.n
        if increment > 0:
            self.pbar.update(increment)

    def close(self):
        self.pbar.close()


# === Глобальные переменные для worker'ов ===
WORKER_OCR_LANG = "eng"
WORKER_TESSERACT_CMD = None
WORKER_LOG_FILE = "pdf_search.log"
LOG_FILE_PATH = None


def _worker_init(ocr_lang: str, tesseract_cmd: Optional[str], log_file: str):
    global WORKER_OCR_LANG, WORKER_TESSERACT_CMD, WORKER_LOG_FILE
    WORKER_OCR_LANG = ocr_lang
    WORKER_TESSERACT_CMD = tesseract_cmd
    WORKER_LOG_FILE = log_file

    if WORKER_TESSERACT_CMD:
        pytesseract.pytesseract.tesseract_cmd = WORKER_TESSERACT_CMD

    # Настройка логгера в каждом worker'е
    logger = logging.getLogger("worker")
    if not logger.handlers:
        handler = logging.FileHandler(log_file, encoding="utf-8")
        formatter = logging.Formatter(
            "%(asctime)s | %(levelname)-8s | Worker | %(message)s"
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)


def preprocess_for_ocr(img: Image.Image) -> Image.Image:
    """
    Улучшает изображение для OCR:
    - конвертирует в grayscale,
    - повышает контраст,
    - бинаризует (чёрный текст на белом фоне),
    - удаляет шум.
    """
    # 1. Grayscale
    img = img.convert("L")

    # 2. Повышение контраста
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)

    # 3. Бинаризация
    threshold = 140
    img = img.point(lambda x: 0 if x < threshold else 255, "1")

    # 4. Удаление шума
    img = img.filter(ImageFilter.MedianFilter(size=3))

    return img


def _extract_text_worker(
    pdf_path: str,
    max_pages_total: Optional[int],
) -> Optional[Dict[str, Union[str, int, float]]]:  # <-- изменён тип возврата
    """Worker function with minimized overhead, single-file-read, and OCR logging."""
    path_obj = Path(pdf_path)
    path_str = str(path_obj.resolve())
    logger = logging.getLogger("worker")

    try:
        # Получаем mtime до чтения содержимого
        mtime = path_obj.stat().st_mtime

        # === ЕДИНСТВЕННОЕ ЧТЕНИЕ ФАЙЛА ===
        with open(path_obj, "rb") as f:
            file_bytes = f.read()
        file_hash = hashlib.md5(file_bytes).hexdigest()
        # =================================

        # Извлечение текста через PyMuPDF из байтов
        def extract_with_fitz(buf: bytes, limit: Optional[int]) -> str:
            try:
                doc = fitz.open(stream=buf, filetype="pdf")
                total_pages = len(doc)
                page_limit = total_pages if limit is None else min(total_pages, limit)
                parts = []
                for i in range(page_limit):
                    txt = doc[i].get_text().strip()
                    if txt:
                        parts.append(txt)
                doc.close()
                return "\n".join(parts)
            except Exception as e:
                logger.warning(f"PyMuPDF error on {path_obj.name}: {e}")
                return ""

        # OCR с теми же байтами
        def perform_ocr(buf: bytes, limit: Optional[int], lang: str) -> str:
            try:
                doc = fitz.open(stream=buf, filetype="pdf")
                total_pages = len(doc)
                page_limit = total_pages if limit is None else min(total_pages, limit)
                parts = []
                for i in range(page_limit):
                    pix = doc[i].get_pixmap(dpi=300)  # Повышен DPI для лучшего качества
                    img_data = pix.tobytes(
                        "png"
                    )  # PNG вместо JPG для избежания артефактов сжатия
                    img = Image.open(io.BytesIO(img_data))
                    # === ПРЕДОБРАБОТКА ===
                    clean_img = preprocess_for_ocr(img)
                    txt = pytesseract.image_to_string(
                        clean_img, lang=lang, config="--psm 1"
                    )
                    img.close()
                    parts.append(txt)
                doc.close()
                return "\n".join(parts).strip()
            except Exception as e:
                logger.error(f"OCR error on {path_obj.name}: {e}")
                return ""

        # Основная логика: всегда извлекаем текст, OCR только если текст пуст/слишком короткий
        text = extract_with_fitz(file_bytes, max_pages_total)
        final_content = text

        # Применяем OCR, если текст отсутствует или содержит менее 20 символов (возможно, скан)
        if not text.strip() or len(text.strip()) < 20:
            logger.info(
                f"Starting OCR on {path_obj.name} (text too short: {len(text)} chars)..."
            )
            ocr_text = perform_ocr(file_bytes, max_pages_total, WORKER_OCR_LANG)
            if ocr_text:
                logger.info(
                    f"OCR completed ({len(ocr_text)} chars) for {path_obj.name}"
                )
                final_content = ocr_text
            else:
                # Даже если OCR не дал результата, оставляем исходный текст (может быть пустым)
                final_content = text

        if final_content.strip():
            return {
                "path": path_str,
                "hash": file_hash,
                "text": final_content.strip(),
                "mtime": mtime,  # <-- добавлено
            }
        else:
            # Fallback: используем имя файла, если ничего не извлечено
            return {
                "path": path_str,
                "hash": file_hash,
                "text": path_obj.stem,
                "mtime": mtime,  # <-- добавлено
            }

    except Exception as e:
        logger.error(f"Worker failed on {pdf_path}: {e}")
        try:
            fallback_hash = hashlib.md5(path_obj.read_bytes()).hexdigest()
            fallback_mtime = path_obj.stat().st_mtime
            return {
                "path": path_str,
                "hash": fallback_hash,
                "text": path_obj.stem,
                "mtime": fallback_mtime,  # <-- добавлено
            }
        except Exception as e2:
            logger.error(f"Fallback failed: {e2}")
            return None


class PDFSearchEngine:
    def __init__(self, config_path: str = "data/config.yaml"):
        self.logger = logging.getLogger(self.__class__.__name__)
        self.config = self._load_config(config_path)

        paths = self.config["paths"]
        ocr_config = self.config.get("ocr", {})
        indexing_config = self.config.get("indexing", {})

        self.docs_dir = Path(paths["docs_dir"])
        self.index_dir = Path(paths["index_dir"])
        self.ocr_lang = ocr_config.get("lang", "eng")
        self.max_pages_total = ocr_config.get(
            "max_pages_total", None
        )  # Может быть None
        self.months_back = indexing_config.get("months_back", None)
        self.skip_dirs: Set[str] = set(indexing_config.get("skip_dirs", []))

        self.index_dir.mkdir(parents=True, exist_ok=True)
        self.tesseract_cmd = self._setup_tesseract()
        self.ix = self._initialize_index()

    def _load_config(self, config_path: str) -> Dict[str, Any]:
        try:
            with open(config_path, encoding="utf-8") as f:
                return yaml.safe_load(f)
        except FileNotFoundError as e:
            self.logger.error(f"Config file not found: {config_path}")
            if not Path(config_path).exists():
                self.logger.error("Creating default config template...")
                self._create_default_config(config_path)
                raise ConfigError(
                    f"Config file not found and needs manual creation: {config_path}"
                ) from e
        except yaml.YAMLError as e:
            self.logger.error(f"YAML parsing error: {e}")
            raise ConfigError(f"YAML parsing error: {e}") from e

    def _create_default_config(self, config_path: str) -> None:
        default_config = {
            "paths": {"docs_dir": "./pdfs", "index_dir": "./index"},
            "ocr": {"lang": "rus+eng", "max_pages_total": 100},
            "indexing": {
                "months_back": 3,
                "skip_dirs": ["ДОВЕРЕННОСТИ"],
            },
        }
        with open(config_path, "w", encoding="utf-8") as f:
            yaml.dump(default_config, f, default_flow_style=False, allow_unicode=True)

    def _check_tesseract_installed(self) -> bool:
        try:
            result = subprocess.run(
                ["tesseract", "--version"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
            return result.returncode == 0
        except FileNotFoundError:
            return False

    def _setup_tesseract(self) -> Optional[str]:
        possible_paths = [
            Path(__file__).parent.parent / "static" / "tesseract" / "tesseract.exe",
            Path("/usr/bin/tesseract"),
            Path("/opt/homebrew/bin/tesseract"),
            Path("/usr/local/bin/tesseract"),
        ]

        for path in possible_paths:
            if path.exists():
                self.logger.info(f"Tesseract found at: {path}")
                return str(path)

        script_dir = Path(__file__).parent.resolve()
        tesseract_path = script_dir.parent / "static" / "tesseract" / "tesseract.exe"

        if tesseract_path.exists():
            self.logger.info(f"Tesseract found: {tesseract_path}")
            return str(tesseract_path)
        elif not self._check_tesseract_installed():
            raise RuntimeError(
                f"Tesseract not installed. Expected path: {tesseract_path}"
            )
        return None

    def _initialize_index(self):
        # Добавлено поле mtime
        schema = Schema(
            path=ID(stored=True, unique=True),
            content=TEXT(stored=False),
            hash=ID(stored=True),
            mtime=NUMERIC(stored=True, numtype=float),  # <-- новое поле
        )

        if not index.exists_in(self.index_dir):
            return index.create_in(self.index_dir, schema)
        else:
            return index.open_dir(self.index_dir)

    def _should_skip_directory(self, pdf_path: Path) -> bool:
        try:
            rel_path = pdf_path.parent.relative_to(self.docs_dir)
        except ValueError:
            return False
        for part in rel_path.parts:
            if part in self.skip_dirs:
                return True
        return False

    def _get_filtered_pdf_paths(self) -> List[Path]:
        """Фильтрация PDF с учётом skip_dirs и months_back ДО запуска worker'ов."""
        all_pdfs = list(self.docs_dir.rglob("*.pdf"))
        filtered_pdfs = [p for p in all_pdfs if not self._should_skip_directory(p)]

        if self.months_back is not None:
            import calendar
            from datetime import datetime

            now = datetime.now()

            def subtract_months(dt, months):
                year = dt.year - (months // 12)
                month = dt.month - (months % 12)
                if month <= 0:
                    year -= 1
                    month += 12
                day = min(dt.day, calendar.monthrange(year, month)[1])
                return dt.replace(year=year, month=month, day=day)

            cutoff_date = subtract_months(now, self.months_back)
            final_pdfs = []
            for pdf_path in filtered_pdfs:
                try:
                    mtime = datetime.fromtimestamp(pdf_path.stat().st_mtime)
                    if mtime >= cutoff_date:
                        final_pdfs.append(pdf_path)
                except OSError:
                    continue
            return final_pdfs
        else:
            return filtered_pdfs

    def index_documents(
        self, progress_bar: Optional[ConsoleProgressBar] = None
    ) -> None:
        # Загружаем существующие документы из индекса
        existing_docs = {}
        try:
            with self.ix.searcher() as searcher:
                for hit in searcher.documents():
                    existing_docs[hit["path"]] = {
                        "hash": hit.get("hash", ""),
                        "mtime": hit.get("mtime", 0),
                    }
        except Exception as e:
            self.logger.debug(f"Document loading error: {e}")

        pdf_files = self._get_filtered_pdf_paths()
        total_files = len(pdf_files)

        # Предварительная фильтрация: определяем, какие файлы нужно обрабатывать
        to_process = []
        skipped_count = 0
        for p in pdf_files:
            path_str = str(p.resolve())
            try:
                current_mtime = p.stat().st_mtime
            except OSError:
                # Если не можем получить mtime, обрабатываем как новый
                to_process.append(p)
                continue

            if path_str in existing_docs:
                stored = existing_docs[path_str]
                # Пропускаем, если mtime и хеш совпадают
                if (
                    abs(stored["mtime"] - current_mtime) < 1e-6
                    and stored["hash"] == hashlib.md5(p.read_bytes()).hexdigest()
                ):
                    skipped_count += 1
                    continue
            # Если файл новый или изменился — добавляем в обработку
            to_process.append(p)

        if progress_bar:
            progress_bar.pbar.total = total_files
            progress_bar.update(
                0, f"Parallel text extraction ({len(to_process)} new/changed files)..."
            )

        num_workers = min(multiprocessing.cpu_count(), 8)
        writer = self.ix.writer()
        processed_count = 0
        commit_batch_size = 50

        # Обрабатываем только новые/изменённые файлы
        if to_process:
            with ProcessPoolExecutor(
                max_workers=num_workers,
                initializer=_worker_init,
                initargs=(self.ocr_lang, self.tesseract_cmd, LOG_FILE_PATH),
            ) as executor:
                futures = {
                    executor.submit(
                        _extract_text_worker,
                        str(p.resolve()),
                        self.max_pages_total,
                    ): p
                    for p in to_process
                }

                completed = 0
                for future in as_completed(futures):
                    res = future.result()
                    if res:
                        path_str = res["path"]
                        current_hash = res["hash"]
                        current_mtime = res["mtime"]

                        # Обновляем документ в индексе
                        writer.update_document(
                            path=path_str,
                            content=res["text"],
                            hash=current_hash,
                            mtime=current_mtime,
                        )
                        processed_count += 1

                        if processed_count % commit_batch_size == 0:
                            writer.commit(merge=False)
                            writer = self.ix.writer()

                    completed += 1
                    if progress_bar:
                        progress_bar.update(
                            completed
                            + skipped_count,  # Общее количество обработанных файлов
                            f"Extracted: {completed}/{len(to_process)}",
                        )

        writer.commit(optimize=True)

        self.logger.info(
            f"Indexing complete: {processed_count} indexed, {skipped_count} skipped"
        )

        if progress_bar:
            progress_bar.update(total_files, "Indexing finished")

    def search(self, query_str: str, limit: int = 10) -> List[Tuple[str, float]]:
        try:
            with self.ix.searcher() as searcher:
                parser = QueryParser("content", self.ix.schema)
                query = parser.parse(query_str)
                results = searcher.search(query, limit=limit)
                return [(hit["path"], hit.score) for hit in results]
        except Exception as e:
            self.logger.error(f"Search error: {e}")
            return []


def process_excel_with_pdf_links(
    pdf_search_engine: PDFSearchEngine,
    excel_path: str,
    progress_bar: Optional[ConsoleProgressBar] = None,
    max_files_per_order: int = 5,
) -> Optional[str]:
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if "Код заказа" not in headers:
        raise ValueError('Column "Код заказа" not found!')

    code_col_idx = headers.index("Код заказа") + 1

    # Определяем максимальный индекс столбца, чтобы добавить новые столбцы в конец
    max_col_idx = len(headers)

    # Добавляем столбец "Номер реестра" после последнего существующего столбца
    registry_col_idx = max_col_idx + 1
    ws.cell(row=1, column=registry_col_idx, value="Номер реестра")

    # Добавляем столбцы для файлов после столбца "Номер реестра"
    file_cols = []
    for i in range(1, max_files_per_order + 1):
        col_idx = registry_col_idx + i
        ws.cell(row=1, column=col_idx, value=f"Файл {i}")
        file_cols.append(col_idx)

    total_rows = ws.max_row - 1

    if progress_bar:
        progress_bar.pbar.total = total_rows
        progress_bar.update(0, f"Processing Excel: {total_rows} records...")

    for row in range(2, ws.max_row + 1):
        code = ws.cell(row=row, column=code_col_idx).value
        if not code:
            continue

        # Ищем номер заказа в формате ЗДИ-XXXXXX или ЗДИ XXXXXX
        match = re.search(r"ЗДИ[-\s]*(\d+)", str(code))
        if not match:
            continue

        order_number = match.group(1)
        logging.info(f"Searching for order: {order_number}, row {row}")

        # Поиск файлов по номеру заказа для получения файла с номером реестра
        order_results = pdf_search_engine.search(order_number, limit=10)

        registry_number = None

        # Ищем файл, содержащий номер реестра
        for path, _score in order_results:
            try:
                # Читаем содержимое файла для поиска номера реестра
                with open(path, "rb") as f:
                    file_bytes = f.read()

                # Используем PyMuPDF для извлечения текста
                doc = fitz.open(stream=file_bytes, filetype="pdf")
                full_text = ""
                for i in range(min(5, len(doc))):  # Читаем первые 5 страниц
                    full_text += doc[i].get_text()
                doc.close()

                # Ищем номер реестра в формате РС-XXXXXXX
                registry_match = re.search(r"РС-\d+", full_text)
                if registry_match:
                    registry_number = registry_match.group()
                    logging.info(
                        f"Found registry number {registry_number} in file {path}"
                    )
                    break

            except Exception as e:
                logging.warning(f"Error reading file {path}: {e}")
                continue

        # Если номер реестра найден, ищем все файлы с этим номером
        if registry_number:
            # Записываем номер реестра в соответствующий столбец
            ws.cell(row=row, column=registry_col_idx, value=registry_number)

            # Ищем все файлы, содержащие номер реестра
            all_results = pdf_search_engine.search(
                registry_number, limit=max_files_per_order
            )

            # Записываем найденные файлы в столбцы
            for i, (path, _score) in enumerate(all_results):
                if i < len(file_cols):
                    cell = ws.cell(row=row, column=file_cols[i], value=path)
                    cell.hyperlink = path
                    if hasattr(cell, "style"):
                        cell.style = "Hyperlink"
        else:
            # Если номер реестра не найден, просто ищем файлы по номеру заказа
            logging.info(
                f"No registry number found for order {order_number}, searching by order number"
            )
            all_results = pdf_search_engine.search(
                order_number, limit=max_files_per_order
            )

            # Записываем номер заказа в столбец "Номер реестра" (если не нашли настоящий номер реестра)
            ws.cell(row=row, column=registry_col_idx, value=f"ЗДИ-{order_number}")

            # Записываем найденные файлы в столбцы
            for i, (path, _score) in enumerate(all_results):
                if i < len(file_cols):
                    cell = ws.cell(row=row, column=file_cols[i], value=path)
                    cell.hyperlink = path
                    if hasattr(cell, "style"):
                        cell.style = "Hyperlink"

        if progress_bar:
            progress_bar.update(row - 1, f"Processed: {row - 1}/{total_rows}")

    temp_path = tempfile.mktemp(suffix="_PDF_links.xlsx")
    wb.save(temp_path)
    logging.info(f"Created temp file: {temp_path}")
    return temp_path


def select_excel_file() -> Optional[str]:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    file_path = filedialog.askopenfilename(
        title="Select Excel file with order registry",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return file_path


def show_system_notification(title: str, message: str) -> None:
    try:
        notification.notify(title=title, message=message, timeout=5)
    except Exception as e:
        logging.warning(f"Notification error: {e}")


def open_file_with_default_app(filepath: str) -> None:
    try:
        if sys.platform == "win32":
            os.startfile(filepath)
        elif sys.platform == "darwin":
            subprocess.run(["open", filepath], check=True)
        else:
            subprocess.run(["xdg-open", filepath], check=True)
    except subprocess.CalledProcessError as e:
        logging.error(f"Failed to open file {filepath}: {e}")


def main() -> None:
    logger = setup_logging()
    logger.info("=== Starting PDF Search ===")
    try:
        engine = PDFSearchEngine(config_path="data/config.yaml")

        response = input("Perform document indexing? (y/n): ").strip().lower()
        if response in ["y", "yes", "да", "д"]:
            progress = ConsoleProgressBar("Indexing", total=100)
            engine.index_documents(progress_bar=progress)
            progress.close()
        else:
            logger.info("Indexing skipped")

        excel_path = select_excel_file()
        if not excel_path:
            logger.info("No file selected")
            return

        progress = ConsoleProgressBar("Processing Excel", total=100)
        output_file = process_excel_with_pdf_links(
            engine, excel_path, progress_bar=progress, max_files_per_order=5
        )
        progress.close()

        if output_file:
            open_file_with_default_app(output_file)
            show_system_notification(
                "PDF Search", "Done! Excel file with hyperlinks opened."
            )
            logger.info("Completed successfully")
        else:
            show_system_notification("PDF Search", "Operation cancelled.")
            logger.info("Operation cancelled")

    except Exception as e:
        logger.exception(f"Critical error: {e}")
        show_system_notification("PDF Search", "An error occurred! See log.")
        sys.exit(1)


if __name__ == "__main__":
    main()
